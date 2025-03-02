# ReadMe
# Date: March 1, 2025
#
# Python Libraries Installation:
# Install the required Python libraries using pip:
#    pip install selenium webdriver-manager openpyxl
#
# Additional Tools:
# - Google Chrome web driver
#
# URL:
# - The script scrapes data from the contest standings page:
#   please put your desired contest standings link (bapsoj / toph)  in the url variable below
#
# Instructions:
# 1. Run the script to generate the Excel file with the contest standings.
# 2. The output file will be saved in xlsx format.


from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
import time
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from collections import defaultdict

# Set up Chrome options (Headless mode for performance)
options = Options()
options.add_argument("--headless")
options.add_argument("--disable-gpu")
options.add_argument("--no-sandbox")
options.add_argument("--disable-dev-shm-usage")

# Initialize Chrome driver
driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

# URL of the contest standings page
url = 'https://toph.co/c/mtb-presents-aust-inter-university-2025/standings'

# Open the webpage
driver.get(url)

# Wait for the page to load completely
time.sleep(5)

# Find the table containing the standings
table = driver.find_element(By.TAG_NAME, 'table')

# Extract table rows
university_map = defaultdict(list)

for tr in table.find_elements(By.TAG_NAME, 'tr')[1:]:  # Skip the header row
    cells = tr.find_elements(By.TAG_NAME, 'td')

    if len(cells) < 3:
        continue  # Skip rows with insufficient data

    rank = cells[0].text.strip()
    team_university_info = cells[1].text.strip().split("\n")
    solved_info = cells[2].text.strip().split("\n")

    if len(team_university_info) < 2 or len(solved_info) < 1:
        continue  # Skip invalid rows

    team_name = team_university_info[0]  # Team name
    university = team_university_info[1]  # University name
    solved = solved_info[0]  # Number of problems solved

    try:
        rank = int(rank)
        solved = int(solved)
    except ValueError:
        continue  # Skip rows with invalid rank or solved count

    # Store data in the university map
    university_map[university].append((rank, solved, team_name))

# Close the browser
driver.quit()

# Sort universities by their best (lowest) team rank
sorted_universities = sorted(university_map.keys(), key=lambda uni: min(team[0] for team in university_map[uni]))

# Create a new Excel workbook and select the active sheet
wb = Workbook()
ws = wb.active

# Write header row
header = ["University Rank", "University Name"]
max_teams = max(len(teams) for teams in university_map.values())
for i in range(max_teams):
    header.append(f"Team {i+1}")
ws.append(header)

# Apply formatting to the header row
bold_font = Font(bold=True)
for col in range(1, len(header) + 1):
    cell = ws.cell(row=1, column=col)
    cell.font = bold_font
    cell.alignment = Alignment(horizontal="center", vertical="center")

# Function to format rank suffix
def format_rank(rank):
    if rank == 1:
        return "1st"
    elif rank == 2:
        return "2nd"
    elif rank == 3:
        return "3rd"
    else:
        return f"{rank}th"

# Write data rows
row_index = 2  # Start from row 2 (after header)
for uni_rank, university in enumerate(sorted_universities, start=1):
    teams = sorted(university_map[university], key=lambda x: (x[0], -x[1]))  # Sort teams by (rank ASC, solved DESC)
    
    # Write University Rank and University Name (merged across 3 rows)
    ws.cell(row=row_index, column=1, value=uni_rank)
    ws.cell(row=row_index, column=2, value=university)
    ws.merge_cells(start_row=row_index, start_column=1, end_row=row_index + 2, end_column=1)
    ws.merge_cells(start_row=row_index, start_column=2, end_row=row_index + 2, end_column=2)
    
    # Write team data (3 rows per team)
    for team_idx, team in enumerate(teams):
        rank, solved, team_name = team
        col = 3 + team_idx  # Each team occupies one column
        ws.cell(row=row_index, column=col, value=format_rank(rank))
        ws.cell(row=row_index + 1, column=col, value=f"{solved} Solved")
        ws.cell(row=row_index + 2, column=col, value=team_name)
    
    # Fill in blanks if there are fewer teams than the maximum
    for team_idx in range(len(teams), max_teams):
        col = 3 + team_idx
        ws.cell(row=row_index, column=col, value="")
        ws.cell(row=row_index + 1, column=col, value="")
        ws.cell(row=row_index + 2, column=col, value="")
    
    # Move to the next university (3 rows down)
    row_index += 3

# Apply middle alignment to all cells
for row in ws.iter_rows():
    for cell in row:
        cell.alignment = Alignment(horizontal="center", vertical="center")

# Auto-adjust column widths
for col in ws.columns:
    max_length = 0
    column = col[0].column_letter  # Get the column name
    for cell in col:
        try:
            if len(str(cell.value)) > max_length:
                max_length = len(str(cell.value))
        except:
            pass
    adjusted_width = (max_length + 2) * 1.2  # Add some padding
    ws.column_dimensions[column].width = adjusted_width

# Freeze the header row
ws.freeze_panes = "A2"

# Save the Excel file
output_file = "university_based_standings_toph.xlsx"
wb.save(output_file)

print(f"University-based standings saved to '{output_file}'")